from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Form
from fastapi.encoders import jsonable_encoder
from sqlalchemy.orm import Session
import pandas as pd
from io import BytesIO
import unicodedata
from openpyxl import load_workbook

from app.database import get_db
from app.models import Eje
from datetime import date

router = APIRouter()


@router.get("/")
def listar_eje(db: Session = Depends(get_db)):
    registros = db.query(Eje).all()
    return jsonable_encoder(registros)


@router.post("/upload")
async def upload_eje(
    fecha_corte: date | None = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser Excel (.xlsx o .xls)")

    contents = await file.read()
    # Leemos sin encabezados porque hay filas iniciales y múltiples bloques
    df = pd.read_excel(BytesIO(contents), header=None)

    # Cargar también con openpyxl para detectar filas en negrita
    wb = load_workbook(BytesIO(contents), data_only=True)
    ws = wb.active
    bold_rows = set()
    for row in ws.iter_rows():
        if any(cell.font and cell.font.bold for cell in row):
            # openpyxl es 1-based, pandas 0-based
            bold_rows.add(row[0].row - 1)

    def _normalize(value) -> str:
        if pd.isna(value):
            return ""
        # Texto base
        s = str(value).replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
        s = s.strip().lower()
        # Quitar tildes/acentos
        s = "".join(
            c
            for c in unicodedata.normalize("NFD", s)
            if unicodedata.category(c) != "Mn"
        )
        # Colapsar espacios múltiples
        s = " ".join(s.split())
        return s

    # Mapeo de nombres de columna (normalizados) a campos del modelo
    header_name_map = {
        "tipo": "tipo",
        "cta": "cta",
        "subc": "subc",
        "objg": "objg",
        "ord": "ord",
        "sord": "sord",
        "item": "item",
        "sitem": "sitem",
        "concepto": "concepto",
        "fuente": "fuente",
        "situacion": "situacion",
        "rec": "rec",
        "recurso": "recurso",
        "apropiacion vigente dep.gsto.": "apropiacion_vigente_dep_gsto",
        "total cdp dep.gstos": "total_cdp_dep_gstos",
        "apropiacion disponible dep.gsto.": "apropiacion_disponible_dep_gsto",
        "total cdp modificacion dep.gstos": "total_cdp_modificacion_dep_gstos",
        "total compromiso dep.gstos": "total_compromiso_dep_gstos",
        "cdp por comprometer dep.gstos": "cdp_por_comprometer_dep_gstos",
        "total obligaciones dep.gstos": "total_obligaciones_dep_gstos",
        "compromiso por obligar dep.gstos": "compromiso_por_obligar_dep_gstos",
        "total ordenes de pago dep.gstos": "total_ordenes_pago_dep_gstos",
        "obligaciones por ordenar dep.gstos": "obligaciones_por_ordenar_dep_gstos",
        "pagos dep.gstos": "pagos_dep_gstos",
        "ordenes de pago por pagar dep.gstos": "ordenes_pago_por_pagar_dep_gstos",
        "total reintegros dep.gstos": "total_reintegros_dep_gstos",
    }

    def _nan_to_none(v):
        return None if pd.isna(v) else v

    # Convertir números en formato local (ej. "20.280.000,00") a float/decimal
    def _parse_number(v):
        if pd.isna(v):
            return None
        if isinstance(v, (int, float)):
            return v
        s = str(v).replace("\xa0", " ").strip()
        if not s:
            return None
        # Quitar separadores de miles "." y usar "," como separador decimal
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None

    records = []
    current_dep = None  # dependencia_de_afectacion_de_gastos actual
    current_cols = None  # mapeo: campo -> índice de columna

    for idx, row in df.iterrows():
        norm_row = [_normalize(v) for v in row]

        # Fila de título de bloque (dependencia de afectación de gastos)
        if norm_row and "dependencia de afectacion de gastos" in norm_row[0]:
            # El encabezado está en la columna A, el valor suele estar en la celda de la derecha (o alguna siguiente no vacía)
            dep_value = None
            for v in row[1:]:
                if not (isinstance(v, float) and pd.isna(v)) and str(v).strip() != "":
                    dep_value = str(v).replace("\xa0", " ").strip()
                    break

            # Si por alguna razón no encontramos valor a la derecha, usamos el propio texto de la celda A como fallback
            current_dep = dep_value or str(row.iloc[0]).replace("\xa0", " ").strip()
            current_cols = None
            continue

        # Fila de encabezados de bloque (donde aparecen TIPO, CTA, etc.)
        if any(n in header_name_map for n in norm_row):
            col_map = {}
            for col_idx, name in enumerate(norm_row):
                field = None
                # Coincidencia directa
                if name in header_name_map:
                    field = header_name_map[name]
                else:
                    # Manejar variantes de "REC" con puntos/espacios ("RE C.", "REC.", etc.)
                    simple = name.replace(" ", "").replace(".", "")
                    if simple == "rec":
                        field = "rec"

                if field:
                    col_map[field] = col_idx
            # Requerimos al menos un subconjunto razonable para considerar que es encabezado
            if {"tipo", "cta", "subc"}.issubset(col_map.keys()):
                current_cols = col_map
            continue

        # Filas de datos: solo si ya tenemos una dependencia y un encabezado definido
        if current_dep and current_cols:
            # Fila vacía -> probable separación entre bloques
            if all((val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "") for val in row):
                continue

            data = {
                "dependecia_de_afectacion_de_gastos": current_dep,
                # Marcar si la fila original estaba en negrita en el Excel
                "es_resumen": idx in bold_rows,
            }
            for field, col_idx in current_cols.items():
                value = row.iloc[col_idx] if col_idx < len(row) else None
                # Campos numéricos que pueden venir con formato "20.280.000,00"
                if field in {
                    "rec",
                    "apropiacion_vigente_dep_gsto",
                    "total_cdp_dep_gstos",
                    "apropiacion_disponible_dep_gsto",
                    "total_cdp_modificacion_dep_gstos",
                    "total_compromiso_dep_gstos",
                    "cdp_por_comprometer_dep_gstos",
                    "total_obligaciones_dep_gstos",
                    "compromiso_por_obligar_dep_gstos",
                    "total_ordenes_pago_dep_gstos",
                    "obligaciones_por_ordenar_dep_gstos",
                    "pagos_dep_gstos",
                    "ordenes_pago_por_pagar_dep_gstos",
                    "total_reintegros_dep_gstos",
                }:
                    parsed = _parse_number(value)
                    # Para rec, que es entero, intentamos castear a int si es posible
                    if field == "rec" and parsed is not None:
                        try:
                            parsed = int(parsed)
                        except (TypeError, ValueError):
                            parsed = None
                    data[field] = parsed
                else:
                    data[field] = _nan_to_none(value)

            data["fecha_corte"] = fecha_corte
            records.append(Eje(**data))

    if not records:
        raise HTTPException(status_code=400, detail="No se encontraron filas válidas en el Excel de EJE")

    db.bulk_save_objects(records)
    db.commit()

    return {"insertados": len(records)}
